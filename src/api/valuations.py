from datetime import UTC, datetime
from io import BytesIO
import re

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.db.session import get_session
from src.models.property import Property
from src.models.valuation import Valuation
from src.schemas.cashflow import (
    AnnualCashFlowSummary,
    KeyMetricsSummary,
    LeaseExpirationEntry,
    RentRollEntry,
    TenantCashFlowDetail,
    TenantRecoveryAuditEntry,
    ValuationRunResponse,
)
from src.schemas.valuation import ValuationCreate, ValuationRead, ValuationUpdate
from src.services.valuation_service import ValuationService

router = APIRouter(prefix="/api/v1", tags=["valuations"])


def _safe_filename(text: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("._")
    return safe or "report"


def _build_rent_roll_workbook(
    *,
    property_name: str,
    valuation_name: str,
    analysis_start_date: str | None,
    rent_roll: list[RentRollEntry],
    area_unit: str,
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise HTTPException(
            status_code=500,
            detail="Excel export dependency missing (openpyxl).",
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Rent Roll"

    ws["A1"] = "OpenDCF Rent Roll"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Property: {property_name}"
    ws["A3"] = f"Valuation: {valuation_name}"
    ws["F2"] = f"Analysis Start: {analysis_start_date or '—'}"
    ws["F3"] = f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"

    base_rent_label = "$/Unit/mo" if area_unit == "unit" else "$/SF/yr"
    headers = [
        "Suite",
        "Space Type",
        "Area",
        "Tenant",
        "Lease Start",
        "Lease End",
        "Lease Type",
        f"Base Rent ({base_rent_label})",
        "Annual Rent",
        "Recovery Type",
        "Escalation",
    ]

    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    start_row = header_row + 1
    for i, row in enumerate(rent_roll):
        r = start_row + i
        lease_start = row.lease_start.isoformat() if row.lease_start else None
        lease_end = row.lease_end.isoformat() if row.lease_end else None
        values = [
            row.suite_name,
            row.space_type,
            float(row.area),
            row.tenant_name or "",
            lease_start,
            lease_end,
            row.lease_type,
            float(row.base_rent_per_unit) if row.base_rent_per_unit is not None else None,
            float(row.annual_rent) if row.annual_rent is not None else None,
            row.recovery_type or "",
            row.escalation_type or "",
        ]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = border
            if c in (3, 8, 9):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    last_row = max(header_row + 1, start_row + len(rent_roll) - 1)

    # Number formats
    for r in range(start_row, last_row + 1):
        ws.cell(r, 3).number_format = "#,##0.00"
        ws.cell(r, 8).number_format = "$#,##0.00"
        ws.cell(r, 9).number_format = "$#,##0"
        ws.cell(r, 5).number_format = "yyyy-mm-dd"
        ws.cell(r, 6).number_format = "yyyy-mm-dd"

    # Column sizing
    widths = {
        1: 16, 2: 14, 3: 12, 4: 24, 5: 12, 6: 12, 7: 12, 8: 16, 9: 14, 10: 14, 11: 14,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{header_row}:K{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"1:{header_row}"
    ws.sheet_view.showGridLines = False

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


@router.post(
    "/properties/{property_id}/valuations",
    response_model=ValuationRead,
    status_code=status.HTTP_201_CREATED,
)
async def create_valuation(
    property_id: str,
    body: ValuationCreate,
    db: AsyncSession = Depends(get_session),
) -> ValuationRead:
    result = await db.execute(select(Property).where(Property.id == property_id))
    if result.scalar_one_or_none() is None:
        raise HTTPException(status_code=404, detail="Property not found")
    valuation = Valuation(property_id=property_id, **body.model_dump())
    db.add(valuation)
    await db.commit()
    await db.refresh(valuation)
    return valuation


@router.get("/properties/{property_id}/valuations", response_model=list[ValuationRead])
async def list_valuations(
    property_id: str,
    db: AsyncSession = Depends(get_session),
) -> list[ValuationRead]:
    result = await db.execute(
        select(Valuation).where(Valuation.property_id == property_id)
    )
    return list(result.scalars().all())


@router.get("/valuations/{valuation_id}", response_model=ValuationRead)
async def get_valuation(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> ValuationRead:
    result = await db.execute(select(Valuation).where(Valuation.id == valuation_id))
    v = result.scalar_one_or_none()
    if v is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    return v


@router.put("/valuations/{valuation_id}", response_model=ValuationRead)
async def update_valuation(
    valuation_id: str,
    body: ValuationUpdate,
    db: AsyncSession = Depends(get_session),
) -> ValuationRead:
    result = await db.execute(select(Valuation).where(Valuation.id == valuation_id))
    v = result.scalar_one_or_none()
    if v is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(v, field, value)
    await db.commit()
    await db.refresh(v)
    return v


@router.delete("/valuations/{valuation_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_valuation(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> None:
    result = await db.execute(select(Valuation).where(Valuation.id == valuation_id))
    v = result.scalar_one_or_none()
    if v is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    await db.delete(v)
    await db.commit()


@router.post("/valuations/{valuation_id}/run", response_model=ValuationRunResponse)
async def run_valuation(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> ValuationRunResponse:
    """
    Execute the DCF engine for this valuation.
    Loads all property data, runs the engine, persists results, returns full output.
    """
    result = await db.execute(select(Valuation).where(Valuation.id == valuation_id))
    if result.scalar_one_or_none() is None:
        raise HTTPException(status_code=404, detail="Valuation not found")

    service = ValuationService(db)
    try:
        return await service.execute_valuation(valuation_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Engine error: {exc!s}")


# ---- Reports ----

@router.get("/valuations/{valuation_id}/reports/cash-flow-summary", response_model=list[AnnualCashFlowSummary])
async def report_cash_flow_summary(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> list[AnnualCashFlowSummary]:
    service = ValuationService(db)
    response = await service.get_results(valuation_id)
    if response is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    return response.annual_cash_flows


@router.get("/valuations/{valuation_id}/reports/rent-roll", response_model=list[RentRollEntry])
async def report_rent_roll(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> list[RentRollEntry]:
    service = ValuationService(db)
    response = await service.get_results(valuation_id)
    if response is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    return response.rent_roll


@router.get("/valuations/{valuation_id}/reports/rent-roll.xlsx")
async def report_rent_roll_excel(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    valuation_result = await db.execute(select(Valuation).where(Valuation.id == valuation_id))
    valuation = valuation_result.scalar_one_or_none()
    if valuation is None:
        raise HTTPException(status_code=404, detail="Valuation not found")

    property_result = await db.execute(select(Property).where(Property.id == valuation.property_id))
    property_ = property_result.scalar_one_or_none()
    if property_ is None:
        raise HTTPException(status_code=404, detail="Property not found")

    service = ValuationService(db)
    response = await service.get_results(valuation_id)
    if response is None:
        raise HTTPException(status_code=404, detail="Valuation not found")

    workbook_bytes = _build_rent_roll_workbook(
        property_name=property_.name,
        valuation_name=valuation.name,
        analysis_start_date=(
            valuation.analysis_start_date_override or property_.analysis_start_date
        ).isoformat(),
        rent_roll=response.rent_roll,
        area_unit=property_.area_unit,
    )
    filename = (
        f"rent_roll_{_safe_filename(property_.name)}_{_safe_filename(valuation.name)}.xlsx"
    )
    return StreamingResponse(
        BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/valuations/{valuation_id}/reports/lease-expirations", response_model=list[LeaseExpirationEntry])
async def report_lease_expirations(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> list[LeaseExpirationEntry]:
    service = ValuationService(db)
    response = await service.get_results(valuation_id)
    if response is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    return response.lease_expiration_schedule


@router.get("/valuations/{valuation_id}/reports/key-metrics", response_model=KeyMetricsSummary)
async def report_key_metrics(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> KeyMetricsSummary:
    service = ValuationService(db)
    response = await service.get_results(valuation_id)
    if response is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    if response.key_metrics is None:
        raise HTTPException(status_code=400, detail="Valuation has not been run yet")
    return response.key_metrics


@router.get("/valuations/{valuation_id}/reports/tenant-detail", response_model=list[TenantCashFlowDetail])
async def report_tenant_detail(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> list[TenantCashFlowDetail]:
    service = ValuationService(db)
    response = await service.get_results(valuation_id)
    if response is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    return response.tenant_cash_flows


@router.get("/valuations/{valuation_id}/reports/recovery-audit", response_model=list[TenantRecoveryAuditEntry])
async def report_recovery_audit(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> list[TenantRecoveryAuditEntry]:
    service = ValuationService(db)
    response = await service.get_results(valuation_id)
    if response is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    return response.recovery_audit


@router.get("/valuations/{valuation_id}/reports/full", response_model=ValuationRunResponse)
async def report_full(
    valuation_id: str,
    db: AsyncSession = Depends(get_session),
) -> ValuationRunResponse:
    service = ValuationService(db)
    response = await service.get_results(valuation_id)
    if response is None:
        raise HTTPException(status_code=404, detail="Valuation not found")
    return response
